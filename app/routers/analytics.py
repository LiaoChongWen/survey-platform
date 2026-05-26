import json
import uuid
import csv as csv_module
import io
from datetime import datetime
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth import get_current_admin
from app.models import Form, Question, Response as ResponseModel, Answer, Dataset
from app.services.chart_data import (
    get_score_distribution, get_question_correct_rate,
    get_survey_question_stats, get_essay_texts,
)
from app.services.wordcloud_gen import generate_wordcloud_b64, get_top_words
from app.services.csv_analysis import read_csv_safe, get_column_info, get_preview, analyze
from app.config import UPLOAD_DIR
from app.schemas import AnalyzeRequest

router = APIRouter()


def _require(request: Request):
    user = get_current_admin(request)
    if not user:
        raise HTTPException(status_code=401, detail="未登入")
    return user


# --- Form analytics ---

@router.get("/forms/{form_id}/stats")
def form_stats(form_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    score_dist = get_score_distribution(form_id, db)
    correct_rates = get_question_correct_rate(form_id, db)
    survey_stats = get_survey_question_stats(form_id, db)
    return {
        "score_distribution": score_dist,
        "question_correct_rates": correct_rates,
        "survey_stats": survey_stats,
    }


@router.get("/forms/{form_id}/wordcloud")
def form_wordcloud(form_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    texts = get_essay_texts(form_id, db)
    if not texts:
        return {"image_b64": "", "top_words": []}
    wc_b64 = generate_wordcloud_b64(texts)
    top = get_top_words(texts)
    return {"image_b64": wc_b64, "top_words": top}


@router.get("/forms/{form_id}/question/{question_id}/wordcloud")
def question_wordcloud(form_id: str, question_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    q = db.query(Question).filter(
        Question.question_id == question_id,
        Question.form_id == form_id
    ).first()
    if not q:
        raise HTTPException(404, "問題不存在")
    answers = db.query(Answer).filter(Answer.question_id == question_id).all()
    texts = []
    for a in answers:
        val = json.loads(a.answer_value or "null")
        if val and isinstance(val, str) and val.strip():
            texts.append(val.strip())
    if not texts:
        return {"image_b64": "", "top_words": []}
    wc_b64 = generate_wordcloud_b64(texts)
    top = get_top_words(texts)
    return {"image_b64": wc_b64, "top_words": top}


@router.get("/forms/{form_id}/responses-table")
def responses_table(form_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    form = db.query(Form).filter(Form.form_id == form_id).first()
    if not form:
        raise HTTPException(404, "問卷不存在")
    responses = db.query(ResponseModel).filter(ResponseModel.form_id == form_id).order_by(ResponseModel.submitted_at).all()
    question_ids = [q.question_id for q in form.questions]
    question_texts = {q.question_id: q.question_text for q in form.questions}

    rows = []
    for r in responses:
        ans_map = {a.question_id: json.loads(a.answer_value or "null") for a in r.answers}
        row = {
            "student_name": r.student_name,
            "student_id": r.student_id,
            "student_class": r.student_class,
            "submitted_at": r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "",
            "total_score": r.total_score,
        }
        for qid in question_ids:
            row[qid] = ans_map.get(qid, "")
        rows.append(row)

    return {
        "headers": ["姓名", "學號", "班級", "送出時間", "總分"] + [question_texts[qid] for qid in question_ids],
        "header_keys": ["student_name", "student_id", "student_class", "submitted_at", "total_score"] + question_ids,
        "rows": rows,
    }


@router.get("/forms/{form_id}/export/csv")
def export_csv(form_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    form = db.query(Form).filter(Form.form_id == form_id).first()
    if not form:
        raise HTTPException(404, "問卷不存在")

    responses = db.query(ResponseModel).filter(ResponseModel.form_id == form_id).all()
    questions = form.questions

    output = io.StringIO()
    writer = csv_module.writer(output)
    headers = ["姓名", "學號", "班級", "送出時間", "總分"] + [q.question_text for q in questions]
    writer.writerow(headers)

    for r in responses:
        ans_map = {a.question_id: json.loads(a.answer_value or "null") for a in r.answers}
        row = [
            r.student_name, r.student_id, r.student_class,
            r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "",
            r.total_score,
        ]
        for q in questions:
            val = ans_map.get(q.question_id, "")
            if isinstance(val, list):
                val = ",".join(str(v) for v in val)
            row.append(val)
        writer.writerow(row)

    output.seek(0)
    filename = f"{form.title}_填答資料.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/forms/{form_id}/export/excel")
def export_excel(form_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    form = db.query(Form).filter(Form.form_id == form_id).first()
    if not form:
        raise HTTPException(404, "問卷不存在")

    responses = db.query(ResponseModel).filter(ResponseModel.form_id == form_id).all()
    questions = form.questions

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "填答資料"

    headers = ["姓名", "學號", "班級", "送出時間", "總分"] + [q.question_text for q in questions]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")

    for r in responses:
        ans_map = {a.question_id: json.loads(a.answer_value or "null") for a in r.answers}
        row = [
            r.student_name, r.student_id, r.student_class,
            r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "",
            r.total_score,
        ]
        for q in questions:
            val = ans_map.get(q.question_id, "")
            if isinstance(val, list):
                val = ",".join(str(v) for v in val)
            row.append(val)
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{form.title}_填答資料.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- CSV analysis ---

@router.post("/csv/upload")
async def upload_csv(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    _require(request)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    file_id = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{file_id}.csv"

    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    try:
        df = read_csv_safe(str(file_path))
    except Exception as e:
        file_path.unlink(missing_ok=True)
        raise HTTPException(400, str(e))

    col_info = get_column_info(df)
    dataset = Dataset(
        dataset_id=file_id,
        dataset_name=file.filename.rsplit(".", 1)[0],
        file_name=file.filename,
        file_path=f"{file_id}.csv",          # 只存檔名，不存絕對路徑
        row_count=len(df),
        column_count=len(df.columns),
        column_info=json.dumps(col_info, ensure_ascii=False),
    )
    db.add(dataset)
    db.commit()
    return {"dataset_id": file_id, "message": "上傳成功", "row_count": len(df), "column_count": len(df.columns)}


def resolve_path(ds) -> str:
    """將 DB 裡的 file_path 解析成當前機器的絕對路徑"""
    p = ds.file_path
    # 若已是完整路徑且存在，直接用
    from pathlib import Path as P
    if P(p).exists():
        return p
    # 否則取檔名，從 UPLOAD_DIR 重新組合
    fname = P(p).name
    return str(UPLOAD_DIR / fname)


@router.get("/csv/{dataset_id}/preview")
def csv_preview(dataset_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    ds = db.query(Dataset).filter(Dataset.dataset_id == dataset_id).first()
    if not ds:
        raise HTTPException(404, "資料集不存在")
    df = read_csv_safe(resolve_path(ds))
    return get_preview(df)


@router.post("/csv/{dataset_id}/analyze")
async def csv_analyze(dataset_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    ds = db.query(Dataset).filter(Dataset.dataset_id == dataset_id).first()
    if not ds:
        raise HTTPException(404, "資料集不存在")
    body = await request.json()
    payload = AnalyzeRequest(**body)
    df = read_csv_safe(resolve_path(ds))

    if payload.chart_type == "wordcloud":
        if payload.x_column not in df.columns:
            return {"error": f"欄位 {payload.x_column} 不存在"}
        texts = df[payload.x_column].dropna().astype(str).tolist()
        texts = [t.strip() for t in texts if t.strip()]
        if not texts:
            return {"error": "欄位無文字資料"}
        wc_b64 = generate_wordcloud_b64(texts)
        top = get_top_words(texts)
        return {
            "type": "wordcloud",
            "image_b64": wc_b64,
            "top_words": top,
            "title": f"{payload.x_column} 文字雲",
        }

    result = analyze(df, payload.chart_type, payload.x_column, payload.y_column)
    return result


@router.delete("/csv/{dataset_id}")
def delete_dataset(dataset_id: str, request: Request, db: Session = Depends(get_db)):
    _require(request)
    ds = db.query(Dataset).filter(Dataset.dataset_id == dataset_id).first()
    if not ds:
        raise HTTPException(404, "資料集不存在")
    try:
        Path(resolve_path(ds)).unlink(missing_ok=True)
    except Exception:
        pass
    db.delete(ds)
    db.commit()
    return {"message": "刪除成功"}


@router.get("/forms/{form_id}/linked-dataset")
def get_linked_dataset(form_id: str, request: Request, db: Session = Depends(get_db)):
    """查詢此問卷是否已有對應的分析 Dataset"""
    _require(request)
    ds = db.query(Dataset).filter(Dataset.form_id == form_id).first()
    if not ds:
        return {"linked": False}
    return {
        "linked": True,
        "dataset_id": ds.dataset_id,
        "dataset_name": ds.dataset_name,
        "row_count": ds.row_count,
        "updated_at": ds.uploaded_at.isoformat() if ds.uploaded_at else None,
    }


@router.post("/forms/{form_id}/import-as-dataset")
def import_form_as_dataset(form_id: str, request: Request, db: Session = Depends(get_db)):
    """將問卷填答資料轉成 CSV 並匯入分析資料集；若已有對應 Dataset 則更新而非重建。"""
    _require(request)
    form = db.query(Form).filter(Form.form_id == form_id).first()
    if not form:
        raise HTTPException(404, "問卷不存在")

    responses = db.query(ResponseModel).filter(ResponseModel.form_id == form_id).all()
    if not responses:
        raise HTTPException(400, "此問卷尚無填答資料")

    questions = form.questions

    # 建立 CSV 內容
    output = io.StringIO()
    writer = csv_module.writer(output)
    headers = ["姓名", "學號", "班級", "送出時間", "總分"] + [q.question_text for q in questions]
    writer.writerow(headers)

    for r in responses:
        ans_map = {a.question_id: json.loads(a.answer_value or "null") for a in r.answers}
        row = [
            r.student_name, r.student_id, r.student_class,
            r.submitted_at.strftime("%Y-%m-%d %H:%M") if r.submitted_at else "",
            r.total_score,
        ]
        for q in questions:
            val = ans_map.get(q.question_id, "")
            if isinstance(val, list):
                val = ",".join(str(v) for v in val)
            elif val is None:
                val = ""
            row.append(val)
        writer.writerow(row)

    csv_content = output.getvalue()

    # 分析欄位資訊（從字串直接解析，不需先寫檔）
    from app.services.csv_analysis import get_column_info
    import pandas as pd
    df = pd.read_csv(io.StringIO(csv_content))
    col_info = get_column_info(df)

    dataset_name = f"{form.title}_填答資料"
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

    # 查詢是否已有對應 Dataset
    existing = db.query(Dataset).filter(Dataset.form_id == form_id).first()

    if existing:
        # 更新：覆寫現有 CSV 檔案，更新 metadata
        file_path = UPLOAD_DIR / existing.file_path
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_content)
        existing.dataset_name = dataset_name
        existing.row_count = len(responses)
        existing.column_count = len(headers)
        existing.column_info = json.dumps(col_info, ensure_ascii=False)
        existing.uploaded_at = datetime.utcnow()
        db.commit()
        return {
            "dataset_id": existing.dataset_id,
            "message": "資料已更新",
            "dataset_name": dataset_name,
            "is_update": True,
        }
    else:
        # 新建：建立 CSV 檔案和 Dataset 記錄
        file_id = str(uuid.uuid4())
        file_path = UPLOAD_DIR / f"{file_id}.csv"
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_content)
        dataset = Dataset(
            dataset_id=file_id,
            dataset_name=dataset_name,
            file_name=f"{dataset_name}.csv",
            file_path=f"{file_id}.csv",
            row_count=len(responses),
            column_count=len(headers),
            column_info=json.dumps(col_info, ensure_ascii=False),
            form_id=form_id,
        )
        db.add(dataset)
        db.commit()
        return {
            "dataset_id": file_id,
            "message": "匯入成功",
            "dataset_name": dataset_name,
            "is_update": False,
        }
